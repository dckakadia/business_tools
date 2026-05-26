import os
import uuid
import pandas as pd
import xml.etree.ElementTree as ET
from flask import Blueprint, request, send_file, render_template, current_app
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

gst_bp = Blueprint('gst', __name__, url_prefix='/gst')

def parse_xml_data(xml_path):
    try:
        import re
        try:
            with open(xml_path, 'r', encoding='utf-16') as f:
                content = f.read()
        except UnicodeError:
            with open(xml_path, 'r', encoding='utf-8', errors='replace') as f:
                content = f.read()

        content = re.sub(r'&#[0-9]+;', '', content)
        content = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x05]', '', content)

        root = ET.fromstring(content)
        xml_data = {}

        for voucher in root.findall('.//VOUCHER'):
            gstin_el = voucher.find('PARTYGSTIN')
            if gstin_el is None or not (gstin_el.text or '').strip():
                continue

            gstin = gstin_el.text.strip()
            inv_no_el = voucher.find('VOUCHERNUMBER')
            if inv_no_el is None:
                inv_no_el = voucher.find('REFERENCE')
            inv_no = (inv_no_el.text or '').strip() if inv_no_el is not None else ''

            party_name_el = voucher.find('PARTYLEDGERNAME')
            party_name = (party_name_el.text or '').strip() if party_name_el is not None else ''

            igst = cgst = sgst = inv_value = 0.0

            for ledger in voucher.findall('LEDGERENTRIES.LIST'):
                lname_el = ledger.find('LEDGERNAME')
                amount_el = ledger.find('AMOUNT')
                isparty_el = ledger.find('ISPARTYLEDGER')

                if lname_el is None or amount_el is None:
                    continue

                lname = (lname_el.text or '').upper()
                try:
                    amt = abs(float(amount_el.text))
                except (ValueError, TypeError):
                    continue

                if isparty_el is not None and (isparty_el.text or '').strip() == 'Yes':
                    inv_value += amt
                elif 'IGST' in lname:
                    igst += amt
                elif 'CGST' in lname:
                    cgst += amt
                elif 'SGST' in lname or 'UTGST' in lname:
                    sgst += amt

            total_tax = igst + cgst + sgst
            key = (gstin, inv_no)
            xml_data[key] = {
                'Party_Name': party_name,
                'IGST': igst,
                'CGST': cgst,
                'SGST': sgst,
                'Total_Tax': total_tax,
                'Invoice_Value': inv_value
            }

        return xml_data
    except Exception as e:
        print(f"Error parsing XML: {e}")
        return {}

def compare_entries(excel_row, xml_data, gstin, tolerance=1.0):
    inv_no = str(excel_row['Inv No']).strip() if pd.notna(excel_row['Inv No']) else ''
    key = (gstin, inv_no)

    if key not in xml_data:
        return None, 'not_in_xml'

    xml_entry = xml_data[key]
    mismatches = []

    try:
        excel_igst = float(excel_row['IGST']) if pd.notna(excel_row['IGST']) else 0.0
        excel_cgst = float(excel_row['CGST']) if pd.notna(excel_row['CGST']) else 0.0
        excel_sgst = float(excel_row['SGST']) if pd.notna(excel_row['SGST']) else 0.0
        excel_t_value = float(excel_row['T Value']) if pd.notna(excel_row['T Value']) else 0.0
        excel_inv_value = float(excel_row['Inv Value']) if pd.notna(excel_row['Inv Value']) else 0.0

        excel_total_tax = excel_igst + excel_cgst + excel_sgst

        igst_match = abs(excel_igst - xml_entry['IGST']) <= tolerance
        cgst_match = abs(excel_cgst - xml_entry['CGST']) <= tolerance
        sgst_match = abs(excel_sgst - xml_entry['SGST']) <= tolerance
        total_tax_match = abs(excel_total_tax - xml_entry['Total_Tax']) <= tolerance
        inv_value_match = abs(excel_inv_value - xml_entry['Invoice_Value']) <= tolerance

        if not igst_match: mismatches.append('IGST')
        if not cgst_match: mismatches.append('CGST')
        if not sgst_match: mismatches.append('SGST')
        if not total_tax_match: mismatches.append('Total Tax')
        if not inv_value_match: mismatches.append('Invoice Value')

        comparison = {
            'GSTIN': gstin,
            'Trade_Name_Excel': excel_row['Trade Name'],
            'Trade_Name_XML': xml_entry['Party_Name'],
            'Inv_No': inv_no,
            'Inv_Date': excel_row['Inv Dt'],
            'T_Value_Excel': excel_t_value,
            'IGST_Excel': excel_igst,
            'IGST_XML': xml_entry['IGST'],
            'IGST_Diff': excel_igst - xml_entry['IGST'],
            'CGST_Excel': excel_cgst,
            'CGST_XML': xml_entry['CGST'],
            'CGST_Diff': excel_cgst - xml_entry['CGST'],
            'SGST_Excel': excel_sgst,
            'SGST_XML': xml_entry['SGST'],
            'SGST_Diff': excel_sgst - xml_entry['SGST'],
            'Total_Tax_Excel': excel_total_tax,
            'Total_Tax_XML': xml_entry['Total_Tax'],
            'Total_Tax_Diff': excel_total_tax - xml_entry['Total_Tax'],
            'Inv_Value_Excel': excel_inv_value,
            'Inv_Value_XML': xml_entry['Invoice_Value'],
            'Inv_Value_Diff': excel_inv_value - xml_entry['Invoice_Value'],
            'Status': 'MATCHED' if not mismatches else 'MISMATCHED',
            'Mismatch_Fields': ', '.join(mismatches) if mismatches else 'All Matched'
        }
        return comparison, ('mismatch' if mismatches else 'matched')

    except Exception as e:
        print(f"Error comparing entry for {gstin} / {inv_no}: {e}")
        return None, 'error'

def apply_formatting(writer, sheet_name, df, is_mismatch_sheet=False, is_matched_sheet=False):
    workbook = writer.book
    worksheet = writer.sheets[sheet_name]
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    total_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    mismatch_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    match_fill = PatternFill(start_color="95E1D3", end_color="95E1D3", fill_type="solid")
    excel_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
    xml_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
    diff_negative_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
    diff_positive_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
    
    header_font = Font(bold=True, color="FFFFFF", size=11)
    total_font = Font(bold=True, size=11)
    bold_font = Font(bold=True)
    
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    for col_num, column in enumerate(df.columns, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
        
        column_letter = get_column_letter(col_num)
        max_length = max(len(str(column)), 12)
        worksheet.column_dimensions[column_letter].width = min(max_length + 2, 30)
    
    last_row = len(df) + 1
    for row_num in range(2, last_row + 1):
        for col_num in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left', vertical='center')
            
            first_cell_value = str(worksheet.cell(row=row_num, column=1).value)
            is_total_row = first_cell_value == 'TOTAL'
            
            if is_total_row:
                cell.fill = total_fill
                cell.font = total_font
                cell.alignment = Alignment(horizontal='right' if col_num > 3 else 'left', vertical='center')
            else:
                column_name = df.columns[col_num - 1]
                if is_mismatch_sheet or is_matched_sheet:
                    if '_Excel' in column_name: cell.fill = excel_fill
                    elif '_XML' in column_name: cell.fill = xml_fill
                    elif '_Diff' in column_name:
                        try:
                            value = float(cell.value) if cell.value else 0
                            if abs(value) > 0.01:
                                cell.fill = diff_negative_fill if value < 0 else diff_positive_fill
                        except: pass
                    
                    if column_name == 'Mismatch_Fields' and is_mismatch_sheet:
                        if cell.value and cell.value != 'All Matched':
                            cell.fill = mismatch_fill
                            cell.font = bold_font
                    
                    if column_name == 'Status':
                        if cell.value == 'MATCHED':
                            cell.fill = match_fill
                            cell.font = bold_font
                        elif cell.value == 'MISMATCHED':
                            cell.fill = mismatch_fill
                            cell.font = bold_font
    
    worksheet.freeze_panes = 'A2'

def add_total_row(df, numeric_columns):
    if df.empty: return df
    total_row = {}
    for col in df.columns:
        if col in numeric_columns: total_row[col] = df[col].sum()
        elif col in ['GSTIN', 'Trade_Name_Excel']: total_row[col] = 'TOTAL'
        else: total_row[col] = ''
    total_df = pd.DataFrame([total_row])
    return pd.concat([df, total_df], ignore_index=True)

@gst_bp.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        excel_file = request.files.get("excel_file")
        xml_file = request.files.get("xml_file")
        
        if not excel_file: return "No Excel file uploaded", 400
        if not xml_file: return "No XML file uploaded", 400
        
        upload_folder = current_app.config['UPLOAD_FOLDER']
        output_folder = current_app.config['OUTPUT_FOLDER']
        
        excel_path = os.path.join(upload_folder, f"{uuid.uuid4().hex}.xlsx")
        excel_file.save(excel_path)
        
        xml_path = os.path.join(upload_folder, f"{uuid.uuid4().hex}.xml")
        xml_file.save(xml_path)
        
        xml_data = parse_xml_data(xml_path)
        
        try:
            df = pd.read_excel(excel_path, sheet_name=0, header=None)
            df = df.iloc[6:].reset_index(drop=True)
            output_df = pd.DataFrame({
                "GSTIN": df.iloc[:, 0], "Trade Name": df.iloc[:, 1], "Inv No": df.iloc[:, 2],
                "Inv Dt": df.iloc[:, 4], "T Value": df.iloc[:, 8], "IGST": df.iloc[:, 9],
                "CGST": df.iloc[:, 10], "SGST": df.iloc[:, 11], "Inv Value": df.iloc[:, 5],
            }).dropna(how="all")
        except Exception as e:
            return f"Error reading Excel file: {e}", 400
        
        unreconciled_list = []
        matched_list = []
        mismatch_list = []
        
        for idx, row in output_df.iterrows():
            gstin = row['GSTIN']
            if pd.notna(gstin):
                gstin_str = str(gstin).strip()
                comparison, category = compare_entries(row, xml_data, gstin_str)
                
                if category == 'not_in_xml':
                    unreconciled_list.append({
                        'GSTIN': gstin_str, 'Trade Name': row['Trade Name'], 'Inv No': row['Inv No'],
                        'Inv Date': row['Inv Dt'], 'T Value': row['T Value'], 'IGST': row['IGST'],
                        'CGST': row['CGST'], 'SGST': row['SGST'], 'Inv Value': row['Inv Value']
                    })
                elif category == 'matched': matched_list.append(comparison)
                elif category == 'mismatch': mismatch_list.append(comparison)
        
        unreconciled_df = add_total_row(pd.DataFrame(unreconciled_list), ['T Value', 'IGST', 'CGST', 'SGST', 'Inv Value']) if unreconciled_list else pd.DataFrame(columns=['GSTIN', 'Trade Name', 'Inv No', 'Inv Date', 'T Value', 'IGST', 'CGST', 'SGST', 'Inv Value'])
        
        numeric_cols = ['T_Value_Excel', 'IGST_Excel', 'IGST_XML', 'IGST_Diff', 'CGST_Excel', 'CGST_XML', 'CGST_Diff', 'SGST_Excel', 'SGST_XML', 'SGST_Diff', 'Total_Tax_Excel', 'Total_Tax_XML', 'Total_Tax_Diff', 'Inv_Value_Excel', 'Inv_Value_XML', 'Inv_Value_Diff']
        matched_df = add_total_row(pd.DataFrame(matched_list), numeric_cols) if matched_list else pd.DataFrame(columns=['GSTIN', 'Trade_Name_Excel', 'Trade_Name_XML', 'Inv_No', 'Inv_Date', 'Status'])
        mismatch_df = add_total_row(pd.DataFrame(mismatch_list), numeric_cols) if mismatch_list else pd.DataFrame(columns=['GSTIN', 'Trade_Name_Excel', 'Trade_Name_XML', 'Inv_No', 'Inv_Date', 'Status'])
        
        output_path = os.path.join(output_folder, "GSTR2B_Complete_Reconciliation_Report.xlsx")
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            summary_data = {
                'Category': ['Total Entries in Excel', 'Unreconciled (Not in XML)', 'Matched (Perfect Match)', 'Mismatched (Amount Differences)', 'Total Reconciled (Matched + Mismatched)'],
                'Count': [len(output_df), len(unreconciled_list), len(matched_list), len(mismatch_list), len(matched_list) + len(mismatch_list)]
            }
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='0_Summary', index=False)
            unreconciled_df.to_excel(writer, sheet_name='1_Unreconciled_Entries', index=False)
            matched_df.to_excel(writer, sheet_name='2_Matched_Entries', index=False)
            mismatch_df.to_excel(writer, sheet_name='3_Mismatched_Entries', index=False)
            
            apply_formatting(writer, '0_Summary', summary_df)
            apply_formatting(writer, '1_Unreconciled_Entries', unreconciled_df)
            apply_formatting(writer, '2_Matched_Entries', matched_df, is_matched_sheet=True)
            apply_formatting(writer, '3_Mismatched_Entries', mismatch_df, is_mismatch_sheet=True)
        
        return send_file(output_path, as_attachment=True, download_name="GSTR2B_Reconciliation_Report.xlsx")
    
    return render_template("gst_reconciliation.html")
